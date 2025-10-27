# app/exporter.py
from __future__ import annotations

import os
import io
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Body, HTTPException
from pydantic import BaseModel

# Composición del logo sobre fondo rojo opaco
from PIL import Image as PILImage

# ==========================
# Router
# ==========================
router = APIRouter(tags=["export"])


# ==========================
# Request schema
# ==========================
class ExportRequest(BaseModel):
    tabla: str
    archivo: str
    metadata: Optional[Dict[str, Any]] = None  # { logo_path, banner_cols_override }


# ==========================
# Utilidades
# ==========================
def _env_hex(name: str, default: str) -> str:
    val = (os.getenv(name, default) or "").strip()
    if val.startswith("#"):
        val = val[1:]
    val = val.upper()
    if len(val) != 6:
        val = default.lstrip("#").upper()
    return val


def _argb(rgb_hex: str) -> str:
    h = rgb_hex.strip().lstrip("#").upper()
    if len(h) != 6:
        raise ValueError("Colors must be 6-hex RGB")
    return "FF" + h  # alpha opaco


def _resolve_logo_path(logo_path: Optional[str]) -> Optional[str]:
    if not logo_path:
        return None
    if not os.path.isabs(logo_path):
        app_dir = os.path.dirname(os.path.abspath(__file__))  # .../app
        cand = os.path.join(os.path.dirname(app_dir), logo_path)  # raíz repo
        if os.path.exists(cand):
            return cand
        cand = os.path.join(app_dir, logo_path)  # relativo a /app
        if os.path.exists(cand):
            return cand
    return logo_path if os.path.exists(logo_path) else None


def _thin_border() -> Border:
    thin = Side(style="thin", color="FFBFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _autofit(ws, min_width: int = 9, max_width: int = 45) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            s = str(cell.value)
            max_len = max(max_len, len(s))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


# --- ancho del banner (A1:Ex) en píxeles para construir el lienzo rojo del mismo ancho ---
def _col_pixels(ws, col_idx: int) -> int:
    letter = get_column_letter(col_idx)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43  # default Excel
    return int(width * 7 + 5)  # aproximación


def _merged_pixels(ws, start_col: int, end_col: int) -> int:
    return sum(_col_pixels(ws, c) for c in range(start_col, end_col + 1))


# ==========================
# Integraciones (IBMi / FTP)
# ==========================
try:
    from app.services.ibmi import get_df_from_ibmi  # (tabla: str) -> pd.DataFrame
except Exception:
    get_df_from_ibmi = None

try:
    from app.services.ftp import upload_file
except Exception:
    upload_file = None


def _fetch_df(table_name: str) -> pd.DataFrame:
    if get_df_from_ibmi is None:
        raise RuntimeError("No está disponible app.services.ibmi.get_df_from_ibmi")
    df = get_df_from_ibmi(table_name)
    if not isinstance(df, pd.DataFrame):
        raise RuntimeError("get_df_from_ibmi no devolvió un DataFrame")
    return df


# ==========================
# Bloques de escritura
# ==========================
def _write_banner_and_logo(
    ws, cols: int, banner_hex: str, logo_path: Optional[str], row: int = 1
) -> None:
    """
    1) Pinta franja roja (A1:Ex).
    2) Genera un PNG temporal **opaco**: lienzo rojo del MISMO ANCHO que la franja,
       y pega encima el logo escalado **solo por altura**.
    3) Inserta esa imagen y ajusta la altura de la fila 1.
    """
    # 1) Franja
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_argb(banner_hex))

    if not logo_path:
        return

    try:
        target_h_px = int(os.getenv("BANNER_LOGO_HEIGHT_PX", "70"))
        left_pad = int(os.getenv("BANNER_LOGO_LEFT_PAD_PX", "12"))

        # 2) Lienzo rojo opaco del ANCHO del merge (A1:Ex)
        banner_w_px = _merged_pixels(ws, 1, cols)
        canvas_w = max(banner_w_px, 100)
        canvas_h = target_h_px
        canvas_rgb = PILImage.new("RGB", (canvas_w, canvas_h), "#" + banner_hex)

        # 3) Cargar logo (puede tener alpha) y escalar por altura
        base = PILImage.open(logo_path).convert("RGBA")
        scale = target_h_px / float(base.height or 1)
        logo_w = max(1, int(base.width * scale))
        logo_h = target_h_px
        logo = base.resize((logo_w, logo_h), PILImage.LANCZOS)

        # 4) Pegar logo sobre lienzo rojo
        x = max(0, min(left_pad, canvas_w - logo_w))
        canvas_rgb.paste(logo, (x, 0), logo)

        # 5) Guardar temporal e insertar
        output_dir = os.getenv("OUTPUT_DIR", "/app/output").strip()
        os.makedirs(output_dir, exist_ok=True)
        tmp_path = os.path.join(output_dir, "__logo_banner_tmp.png")
        canvas_rgb.save(tmp_path, "PNG")  # opaco

        xlimg = XLImage(tmp_path)
        xlimg.anchor = f"A{row}"
        ws.add_image(xlimg)

        # 6) Altura de la fila 1 (px -> pt ≈ 0.75)
        ws.row_dimensions[row].height = target_h_px * 0.75

    except Exception:
        # Si algo falla, al menos queda pintada la franja
        pass


def _write_metadata_from_df(ws, df: pd.DataFrame, start_row: int = 2) -> int:
    meta = df.iloc[:6, :2].fillna("")
    r = start_row
    for _, (label, value) in meta.iterrows():
        c1 = ws.cell(row=r, column=1, value=str(label or ""))
        c2 = ws.cell(row=r, column=2, value=str(value or ""))
        c1.font = Font(bold=True)
        c2.font = Font(bold=True)
        r += 1
    return r


def _write_center_title(
    ws, text: str, cols: int, row: int, font_size: int = 14, color_hex: str = "333333"
) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(size=font_size, bold=True, color=_argb(color_hex))


# ==========================
# Detalle
# ==========================
REQUIRED_HEADERS = ["FECHA", "DESCRIPCION", "DEBITO", "CREDITO", "SALDO CONTABLE"]


def _prepare_detail(df: pd.DataFrame) -> pd.DataFrame:
    if len(df) <= 6:
        raise ValueError("Se requieren al menos 7 filas (6 meta + detalle).")

    detail = df.iloc[6:].reset_index(drop=True).copy()

    # Map por nombre o por posición
    cols = {c.upper().strip(): c for c in detail.columns}
    mapping = {}
    if "FECHA" in cols:
        mapping["FECHA"] = cols["FECHA"]
    if "DESCRIPCION" in cols:
        mapping["DESCRIPCION"] = cols["DESCRIPCION"]
    if "DEBITO" in cols:
        mapping["DEBITO"] = cols["DEBITO"]
    if "CREDITO" in cols:
        mapping["CREDITO"] = cols["CREDITO"]
    if "SALDO CONTABLE" in cols:
        mapping["SALDO CONTABLE"] = cols["SALDO CONTABLE"]
    elif "SALDO" in cols:
        mapping["SALDO CONTABLE"] = cols["SALDO"]

    if len(mapping) < 5:
        if detail.shape[1] < 5:
            raise ValueError("El detalle no tiene al menos 5 columnas.")
        sel = detail.iloc[:, :5]
        sel.columns = REQUIRED_HEADERS
    else:
        sel = detail[list(mapping.values())].copy()
        sel.columns = REQUIRED_HEADERS

    # Normalizaciones
    try:
        sel["FECHA"] = pd.to_datetime(sel["FECHA"], errors="coerce").dt.date
    except Exception:
        pass

    for col in ["DEBITO", "CREDITO", "SALDO CONTABLE"]:
        try:
            sel[col] = (
                sel[col]
                .astype(str)
                .str.replace(".", "", regex=False)
                .str.replace(",", ".", regex=False)
            )
            sel[col] = pd.to_numeric(sel[col], errors="coerce")
        except Exception:
            sel[col] = pd.to_numeric(sel[col], errors="coerce")

    return sel


# ==========================
# Export principal
# ==========================
def export_to_excel(
    tabla: str, archivo: str, metadata: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    1) Banner rojo + logo (sobre lienzo rojo del ancho del merge).
    2) Metadatos (6 filas) en negrita.
    3) Título 'Consulta de movimientos'.
    4) Encabezados fijos.
    5) Detalle desde la 7ª fila.
    6) Freeze panes.
    """
    import time

    t0 = time.time()

    OUTPUT_DIR = os.getenv("OUTPUT_DIR", "/app/output").strip()
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    BANNER_COLOR = _env_hex("BANNER_COLOR", "ED1C24")
    HEADER_GRAY = _env_hex("HEADER_GRAY", "F2F2F2")
    TEXT_DARK = _env_hex("TEXT_DARK", "333333")
    DEFAULT_BANNER_COLS = int(os.getenv("BANNER_COLS", "5") or "5")

    FTP_HOST = os.getenv("FTP_HOST", "").strip()
    FTP_PORT = int(os.getenv("FTP_PORT", "21") or "21")
    FTP_USER = os.getenv("FTP_USER", "").strip()
    FTP_PASSWORD = os.getenv("FTP_PASSWORD", "").strip()
    FTP_DIR = os.getenv("FTP_DIR", "/").strip()
    FTP_TLS = os.getenv("FTP_TLS", "false").strip().lower() == "true"
    WRITE_LOCAL_COPY = os.getenv("WRITE_LOCAL_COPY", "true").strip().lower() != "false"

    meta = metadata or {}
    logo_path = _resolve_logo_path(meta.get("logo_path"))
    banner_cols_override = (
        int(meta["banner_cols_override"]) if meta.get("banner_cols_override") else None
    )

    df = _fetch_df(tabla)
    detail_df = _prepare_detail(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    banner_cols = banner_cols_override or 5 or DEFAULT_BANNER_COLS
    banner_cols = max(banner_cols, 5)

    _write_banner_and_logo(
        ws, cols=banner_cols, banner_hex=BANNER_COLOR, logo_path=logo_path, row=1
    )

    next_row = _write_metadata_from_df(ws, df, start_row=2)
    next_row += 1

    _write_center_title(
        ws,
        "Consulta de movimientos",
        cols=banner_cols,
        row=next_row,
        font_size=14,
        color_hex=TEXT_DARK,
    )
    header_row = next_row + 2

    # Encabezados fijos
    for j, name in enumerate(REQUIRED_HEADERS, start=1):
        c = ws.cell(row=header_row, column=j, value=name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", fgColor=_argb(HEADER_GRAY))
        c.border = _thin_border()

    # Detalle
    r = header_row + 1
    for _, row_vals in detail_df.iterrows():
        # FECHA centrada y con formato
        vfecha = row_vals["FECHA"]
        c = ws.cell(row=r, column=1, value=vfecha)
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()
        if pd.notna(vfecha):
            c.number_format = "yyyy-mm-dd"

        # DESCRIPCION
        c = ws.cell(row=r, column=2, value=row_vals["DESCRIPCION"])
        c.border = _thin_border()

        # Montos
        for j, colname in enumerate(["DEBITO", "CREDITO", "SALDO CONTABLE"], start=3):
            v = row_vals[colname]
            c = ws.cell(row=r, column=j, value=(None if pd.isna(v) else float(v)))
            c.alignment = Alignment(horizontal="right")
            c.border = _thin_border()
            c.number_format = "#,##0.00"

        r += 1

    _autofit(ws, min_width=9, max_width=45)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    local_xlsx = os.path.join(OUTPUT_DIR, f"{archivo}.xlsx")
    if WRITE_LOCAL_COPY:
        wb.save(local_xlsx)
    else:
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

    # FTP opcional
    uploaded_info = None
    if FTP_HOST and FTP_USER and FTP_PASSWORD and upload_file:
        remote_name = archivo
        try:
            if WRITE_LOCAL_COPY:
                uploaded_info = upload_file(
                    local_path=local_xlsx,
                    host=FTP_HOST,
                    port=FTP_PORT,
                    user=FTP_USER,
                    password=FTP_PASSWORD,
                    remote_dir=FTP_DIR,
                    remote_name=remote_name,
                    tls=FTP_TLS,
                )
            else:
                tmp_path = os.path.join(OUTPUT_DIR, f"__tmp_{archivo}.xlsx")
                with open(tmp_path, "wb") as f:
                    f.write(bio.getbuffer())
                uploaded_info = upload_file(
                    local_path=tmp_path,
                    host=FTP_HOST,
                    port=FTP_PORT,
                    user=FTP_USER,
                    password=FTP_PASSWORD,
                    remote_dir=FTP_DIR,
                    remote_name=remote_name,
                    tls=FTP_TLS,
                )
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        except Exception as ex:
            uploaded_info = {"status": "error", "error": str(ex)}

    elapsed_ms = int((time.time() - t0) * 1000)
    return {
        "ok": True,
        "table": tabla,
        "rows": int(detail_df.shape[0]),
        "cols": int(detail_df.shape[1]),
        "uploaded_to": uploaded_info or {"status": "skipped"},
        "local_copy": local_xlsx if WRITE_LOCAL_COPY else None,
        "elapsed_ms": elapsed_ms,
    }


# ==========================
# Endpoint
# ==========================
@router.post("/export")
def export(req: ExportRequest = Body(...)) -> Dict[str, Any]:
    try:
        return export_to_excel(req.tabla, req.archivo, req.metadata)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))